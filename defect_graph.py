import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
import os
import sys
import time
from collections import Counter


def print_step(step_number, message):
    """打印步骤提示"""
    time.sleep(0.25)
    print(f"\n=== {message} ===\n")

def wait_for_enter():
    """等待用户按回车继续"""
    input("按回车键继续...")

def find_header_row(sheet, header_keyword="片号"):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # 确保行至少有3列，然后检查第三列（索引2）
        if row[0] is not None:
            if header_keyword in str(row[0]):
                print(f"检测到表头在第 {row_idx} 行")
                return row_idx
    print("未找到表头，默认返回第 1 行")
    return 1

def is_valid_code(code):
    """检查是否为有效的编码（不以#开头）"""
    if code is None:
        return False

    code_str = str(code)
    # 检查是否以#开头（Excel错误值通常以#开头）
    if code_str.startswith('#'):
        return False

    return True

def load_graph_data(file_path):
    """加载Excel数据，只读取所需的列"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 查找表头行
    header_row = find_header_row(sheet)

    data = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # 检查产品编码是否为有效值（不以#开头）
        slice_code = str(row[0]) if row[0] is not None else None
        if not is_valid_code(slice_code):
            continue

        row_data = {
            "种类": row[1],
            "片号": slice_code,
            "这个缺陷": row[2],
            "哪个缺陷": row[3],
            "就是这个缺陷": row[4],

        }
        data.append(row_data)

    return data

def group_product_codes(product_codes):
    """
    对产品编码进行分组，将亚种编码归类到基础编码
    参数:
    product_codes: 所有产品编码的集合
    返回:
    分组字典: {基础编码: [该组的所有编码]}
    """
    # 找出所有编码的最小长度，这应该是基础编码的长度
    min_length = min(len(code) for code in product_codes) if product_codes else 0

    groups = {}

    for code in product_codes:
        # 提取基础编码部分（前min_length个字符）
        base_code = code[:min_length]

        if base_code not in groups:
            groups[base_code] = []

        groups[base_code].append(code)

    return groups


def analyze_defect_data(graph_data):
    """
    分析缺陷数据，统计各工序缺陷类型的占比
    参数:
    graph_data: 图形数据列表
    返回:
    字典，键为工序名，值为缺陷统计Counter对象
    """
    # 定义要分析的工序列
    process_columns = ["这个缺陷", "哪个缺陷", "就是这个缺陷"]

    defect_stats = {}

    for column in process_columns:
        # 收集该列的所有非空值
        defects = []
        for item in graph_data:
            defect = item.get(column)
            if defect is not None and str(defect).strip() != "":
                defects.append(str(defect).strip())

        # 统计缺陷类型
        defect_counter = Counter(defects)
        defect_stats[column] = defect_counter

    return defect_stats


def create_pie_charts_for_group(workbook, group_name, defect_stats):
    """
    为单个产品编码分组创建饼图并添加到工作簿
    参数:
    workbook: openpyxl工作簿对象
    group_name: 分组名称
    defect_stats: 缺陷统计字典
    """
    # 创建工作表
    sheet_name = f"{group_name}缺陷分析"
    if sheet_name in workbook.sheetnames:
        # 如果已存在，则删除原有工作表
        del workbook[sheet_name]
    ws = workbook.create_sheet(title=sheet_name)

    # 添加表头
    ws['A1'] = "缺陷类型"
    ws['B1'] = "数量"
    ws['C1'] = "占比"

    # 设置表头样式
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # 添加数据
    total_count = sum(sum(counter.values()) for counter in defect_stats.values())
    row = 2

    # 记录每个工序的数据范围
    process_ranges = {}

    for process_name, counter in defect_stats.items():
        # 添加工序标题
        ws[f'A{row}'] = f"{process_name}"
        ws[f'A{row}'].font = Font(bold=True, color="FF0000")
        row += 1

        # 记录工序数据开始行
        process_start = row

        process_total = sum(counter.values())
        for defect_type, count in counter.most_common():
            ws[f'A{row}'] = defect_type
            ws[f'B{row}'] = count
            ws[f'C{row}'] = count / process_total if process_total > 0 else 0
            row += 1

        # 记录工序数据结束行
        process_end = row - 1
        process_ranges[process_name] = (process_start, process_end)

        # 添加空行分隔不同工序
        row += 1

    # 设置百分比格式
    for r in range(2, row):
        if ws[f'C{r}'].value is not None:
            ws[f'C{r}'].number_format = '0.00%'

    # 为每个工序创建饼图
    chart_row = 2
    chart_col = 5  # 从E列开始放置图表

    for process_name, counter in defect_stats.items():
        if not counter:  # 跳过空数据
            continue

        # 获取工序数据范围
        if process_name not in process_ranges:
            continue

        data_start, data_end = process_ranges[process_name]

        # 如果没有有效数据，跳过
        if data_start > data_end:
            continue

        # 创建饼图
        chart = PieChart()
        chart.title = f"{group_name}-{process_name}缺陷分布"

        # 设置数据范围
        labels = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_end)

        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)

        # 设置图表样式
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showVal = False
        chart.dataLabels.showCatName = True

        # 将图表添加到工作表
        ws.add_chart(chart, f"{openpyxl.utils.get_column_letter(chart_col)}{chart_row}")

        # 更新下一个图表位置
        chart_col += 8  # 每个图表占8列宽度
        if chart_col > 20:  # 如果超出Z列，换到下一行
            chart_col = 5
            chart_row += 20

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10



def save_grouped_results_to_excel(grouped_data, output_file="工序缺陷统计.xlsx"):
    """将分组结果保存到Excel文件"""
    # 创建新的工作簿
    workbook = openpyxl.Workbook()

    # 删除默认创建的工作表
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # 为每个分组创建工作表并添加饼图
    for group_name, data in grouped_data.items():
        defect_stats = analyze_defect_data(data)
        create_pie_charts_for_group(workbook, group_name, defect_stats)

    # 创建汇总工作表
    summary_sheet = workbook.create_sheet(title="汇总")
    summary_sheet['A1'] = "产品分组"
    summary_sheet['B1'] = "数据条数"
    summary_sheet['C1'] = "包含的产品编码"

    # 设置表头样式
    for cell in ['A1', 'B1', 'C1']:
        summary_sheet[cell].font = Font(bold=True)
        summary_sheet[cell].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # 添加汇总数据
    row = 2
    for group_name, data in grouped_data.items():
        summary_sheet[f'A{row}'] = group_name
        summary_sheet[f'B{row}'] = len(data)
        summary_sheet[f'C{row}'] = ", ".join(set([item["种类"] for item in data]))
        row += 1

    # 调整列宽
    summary_sheet.column_dimensions['A'].width = 15
    summary_sheet.column_dimensions['B'].width = 10
    summary_sheet.column_dimensions['C'].width = 40

    # 保存文件
    workbook.save(output_file)
    print(f"结果已保存到 {output_file}")


# 主程序
if __name__ == "__main__":
    print_step(1, "工艺缺陷分布")
    print("请确保Excel文件与此程序在同一文件夹中")
    wait_for_enter()

    # 查找Excel文件
    print_step(2, "查找Excel文件")
    excel_files = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            excel_files.append(file)

    if not excel_files:
        print("未找到Excel文件(.xlsx或.xls)")
        print("请将Excel文件放入与此程序相同的文件夹中")
        input("按回车键退出...")
        sys.exit(1)

    # 如果找到多个Excel文件，让用户选择
    if len(excel_files) > 1:
        print_step(3, "发现多个Excel文件，请选择要分析的文件:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")

        while True:
            try:
                choice = int(input("请输入文件编号: "))
                if 1 <= choice <= len(excel_files):
                    file_path = excel_files[choice - 1]
                    break
                else:
                    print("编号无效，请重新输入")
            except ValueError:
                print("请输入有效数字")
    else:
        file_path = excel_files[0]
        print(f"找到文件: {file_path}")
        wait_for_enter()

    print_step(4, "开始分析数据")
    print("正在读取和分析Excel文件...")

    try:
        # 加载数据
        graph_data = load_graph_data(file_path)

        # 提取所有产品编码
        all_product_codes = set(item["种类"] for item in graph_data)
        print(f"找到 {len(all_product_codes)} 个不同的种类")
        groups = group_product_codes(all_product_codes)
        print(f"产品编码被分为 {len(groups)} 个组")

        # 按分组组织数据
        grouped_data = {}
        for group_name, product_codes in groups.items():
            grouped_data[group_name] = [
                item for item in graph_data
                if item["种类"] in product_codes
            ]
            print(f"分组 '{group_name}' 包含 {len(grouped_data[group_name])} 条数据")

        print_step(6, "分析缺陷数据")
        # 保存到Excel
        output_file = "工序缺陷统计.xlsx"
        save_grouped_results_to_excel(grouped_data, output_file)

        print_step(7, "完成")
        print("所有操作已完成!")
        print("您可以在同一文件夹中找到 '工序缺陷统计.xlsx' 文件")
        print("文件中包含了各产品分组的缺陷分析饼图")

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        import traceback

        traceback.print_exc()

    input("按回车键退出程序...")