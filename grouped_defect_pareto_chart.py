import xlsxwriter
from collections import Counter
import os
import sys
import time


def print_step(step_number, message):
    """打印步骤提示"""
    time.sleep(0.25)
    print(f"\n=== {message} ===\n")


def wait_for_enter():
    """等待用户按回车继续"""
    input("按回车键继续...")


def find_header_row(sheet, header_keyword="片号"):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # 确保行至少有3列，然后检查第三列（索引2）
        if len(row) > 2 and row[2] is not None:
            if header_keyword in str(row[2]):
                print(f"检测到表头在第 {row_idx} 行")
                return row_idx
    print("未找到表头，默认返回第 1 行")
    return 1


def is_valid_code(code):
    """检查是否为有效的编码（不以#开头）"""
    if code is None:
        return False

    code_str = str(code)
    # 检查是否以#开头（Excel错误值通常以#开头）
    if code_str.startswith('#'):
        return False

    return True


def load_graph_data(file_path):
    """加载Excel数据，只读取所需的列"""
    import openpyxl  # 仅在数据加载时使用openpyxl

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 查找表头行
    header_row = find_header_row(sheet)

    data = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # 跳过空行
        if row[0] is None:
            continue

        # 检查片号是否为有效值（不以#开头）
        slice_code = str(row[0]) if row[0] is not None else None
        if not is_valid_code(slice_code):
            continue

        row_data = {
            "种类": row[1],
            "片号": slice_code,
            "这个缺陷": row[2],
            "哪个缺陷": row[3],
            "就是这个缺陷": row[4],

        }
        data.append(row_data)

    return data


def group_product_codes(product_codes):
    """
    对种类进行分组，将亚种编码归类到基础编码
    参数:
    product_codes: 所有种类的集合
    返回:
    分组字典: {基础编码: [该组的所有编码]}
    """
    # 找出所有编码的最小长度，这应该是基础编码的长度
    min_length = min(len(code) for code in product_codes) if product_codes else 0

    groups = {}

    for code in product_codes:
        # 提取基础编码部分（前min_length个字符）
        base_code = code[:min_length]

        if base_code not in groups:
            groups[base_code] = []

        groups[base_code].append(code)

    return groups


def analyze_defect_data(graph_data):
    """
    分析缺陷数据，统计各工序缺陷类型的占比
    参数:
    graph_data: 图形数据列表
    返回:
    字典，键为工序名，值为缺陷统计Counter对象
    """
    # 定义要分析的工序列
    process_columns = ["这个缺陷", "哪个缺陷", "就是这个缺陷"]

    defect_stats = {}

    for column in process_columns:
        # 收集该列的所有非空值
        defects = []
        for item in graph_data:
            defect = item.get(column)
            if defect is not None and str(defect).strip() != "":
                defects.append(str(defect).strip())

        # 统计缺陷类型
        defect_counter = Counter(defects)
        defect_stats[column] = defect_counter

    return defect_stats


def create_pareto_chart_for_group(workbook, group_name, defect_stats):
    """
    为单个产品编码分组创建帕累托图并添加到工作簿
    参数:
    workbook: xlsxwriter工作簿对象
    group_name: 分组名称
    defect_stats: 缺陷统计字典
    """
    # 创建工作表
    sheet_name = f"{group_name}缺陷分析"
    worksheet = workbook.add_worksheet(sheet_name)

    # 设置列宽
    worksheet.set_column('A:A', 20)
    worksheet.set_column('B:B', 10)
    worksheet.set_column('C:C', 10)
    worksheet.set_column('D:D', 12)

    # 定义格式
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DDDDDD',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    bold_red_format = workbook.add_format({
        'bold': True,
        'color': 'red'
    })

    percent_format = workbook.add_format({
        'num_format': '0.00%'
    })

    # 添加表头
    worksheet.write('A1', "缺陷类型", header_format)
    worksheet.write('B1', "数量", header_format)
    worksheet.write('C1', "占比", header_format)
    worksheet.write('D1', "累积百分比", header_format)

    # 添加数据
    row = 1  # 从第2行开始（0-indexed）

    # 记录每个工序的数据范围
    process_ranges = {}

    for process_name, counter in defect_stats.items():
        # 添加工序标题
        worksheet.write(row, 0, f"{process_name}", bold_red_format)
        row += 1

        # 记录工序数据开始行
        process_start = row

        # 计算总数并排序（按数量降序）
        process_total = sum(counter.values())
        sorted_defects = counter.most_common()

        # 计算累积百分比
        cumulative_percentage = 0

        for defect_type, count in sorted_defects:
            percentage = count / process_total if process_total > 0 else 0
            cumulative_percentage += percentage

            worksheet.write(row, 0, defect_type)
            worksheet.write(row, 1, count)
            worksheet.write(row, 2, percentage, percent_format)
            worksheet.write(row, 3, cumulative_percentage, percent_format)

            row += 1

        # 记录工序数据结束行
        process_end = row - 1
        process_ranges[process_name] = (process_start, process_end)

        # 添加空行分隔不同工序
        row += 1

    # 为每个工序创建帕累托图
    chart_start_row = row + 2  # 在数据下方空两行开始放置图表
    chart_row_offset = 0

    for process_name, counter in defect_stats.items():
        if not counter:  # 跳过空数据
            continue

        # 获取工序数据范围
        if process_name not in process_ranges:
            continue

        data_start, data_end = process_ranges[process_name]

        # 如果没有有效数据，跳过
        if data_start > data_end:
            continue

        # 创建组合图表（柱状图 + 折线图）
        chart = workbook.add_chart({'type': 'column'})

        # 添加柱状图数据
        chart.add_series({
            'name': f'{process_name}缺陷数量',
            'categories': [sheet_name, data_start, 0, data_end, 0],
            'values': [sheet_name, data_start, 1, data_end, 1],
            'gap': 100,  # 设置柱子之间的间距
        })

        # 添加折线图数据（累积百分比）
        chart2 = workbook.add_chart({'type': 'line'})
        chart2.add_series({
            'name': '累积百分比',
            'categories': [sheet_name, data_start, 0, data_end, 0],
            'values': [sheet_name, data_start, 3, data_end, 3],
            'y2_axis': True,  # 使用次Y轴
            'marker': {'type': 'automatic'},
        })
        chart.combine(chart2)

        # 设置图表标题和样式
        chart.set_title({'name': f"{group_name}-{process_name}帕累托图"})
        chart.set_x_axis({'name': '缺陷类型'})
        chart.set_y_axis({'name': '缺陷数量'})
        chart.set_y2_axis({'name': '累积百分比', 'max': 1.0})

        # 删除图例
        chart.set_legend({'none': True})

        # 设置图表大小（宽度和高度）
        chart.set_size({'width': 800, 'height': 500})

        # 调整绘图区域，使其充分填充图表
        chart.set_plotarea({
            'layout': {
                'x': 0.1,  # 左边距
                'y': 0.1,  # 上边距
                'width': 0.85,  # 宽度占比
                'height': 0.75  # 高度占比
            }
        })

        # 插入图表，从chart_start_row + chart_row_offset开始，列固定在4（E列）
        worksheet.insert_chart(chart_start_row + chart_row_offset, 4, chart, {'x_offset': 0, 'y_offset': 0})

        # 每个图表占25行的高度
        chart_row_offset += 25


def save_grouped_results_to_excel(grouped_data, output_file="工序缺陷帕累托统计.xlsx"):
    """将分组结果保存到Excel文件"""
    # 创建新的工作簿
    workbook = xlsxwriter.Workbook(output_file)

    # 为每个分组创建工作表并添加帕累托图
    for group_name, data in grouped_data.items():
        defect_stats = analyze_defect_data(data)
        create_pareto_chart_for_group(workbook, group_name, defect_stats)

    # 创建汇总工作表
    summary_sheet = workbook.add_worksheet("汇总")

    # 设置列宽
    summary_sheet.set_column('A:A', 15)
    summary_sheet.set_column('B:B', 10)
    summary_sheet.set_column('C:C', 40)

    # 定义表头格式
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DDDDDD',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })

    # 添加表头
    summary_sheet.write('A1', "产品分组", header_format)
    summary_sheet.write('B1', "数据条数", header_format)
    summary_sheet.write('C1', "包含的种类", header_format)

    # 添加汇总数据
    row = 1
    for group_name, data in grouped_data.items():
        summary_sheet.write(row, 0, group_name)
        summary_sheet.write(row, 1, len(data))
        summary_sheet.write(row, 2, ", ".join(set([item["种类"] for item in data])))
        row += 1

    # 保存文件
    workbook.close()
    print(f"结果已保存到 {output_file}")


# 主程序
if __name__ == "__main__":
    print_step(1, "工艺缺陷分布")
    print("请确保Excel文件与此程序在同一文件夹中")
    wait_for_enter()

    # 查找Excel文件
    print_step(2, "查找Excel文件")
    excel_files = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            excel_files.append(file)

    if not excel_files:
        print("未找到Excel文件(.xlsx或.xls)")
        print("请将Excel文件放入与此程序相同的文件夹中")
        input("按回车键退出...")
        sys.exit(1)

    # 如果找到多个Excel文件，让用户选择
    if len(excel_files) > 1:
        print_step(3, "发现多个Excel文件，请选择要分析的文件:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")

        while True:
            try:
                choice = int(input("请输入文件编号: "))
                if 1 <= choice <= len(excel_files):
                    file_path = excel_files[choice - 1]
                    break
                else:
                    print("编号无效，请重新输入")
            except ValueError:
                print("请输入有效数字")
    else:
        file_path = excel_files[0]
        print(f"找到文件: {file_path}")
        wait_for_enter()

    print_step(4, "开始分析数据")
    print("正在读取和分析Excel文件...")

    try:
        # 加载数据
        graph_data = load_graph_data(file_path)

        # 提取所有种类
        all_product_codes = set(item["种类"] for item in graph_data)
        print(f"找到 {len(all_product_codes)} 个不同的种类")
        groups = group_product_codes(all_product_codes)
        print(f"种类被分为 {len(groups)} 个组")

        # 按分组组织数据
        grouped_data = {}
        for group_name, product_codes in groups.items():
            grouped_data[group_name] = [
                item for item in graph_data
                if item["种类"] in product_codes
            ]
            print(f"分组 '{group_name}' 包含 {len(grouped_data[group_name])} 条数据")

        print_step(6, "分析缺陷数据")
        # 保存到Excel
        output_file = "工序缺陷帕累托统计.xlsx"
        save_grouped_results_to_excel(grouped_data, output_file)

        print_step(7, "完成")
        print("所有操作已完成!")
        print("您可以在同一文件夹中找到 '工序缺陷帕累托统计.xlsx' 文件")
        print("文件中包含了各产品分组的缺陷分析帕累托图")

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        import traceback

        traceback.print_exc()

    input("按回车键退出程序...")