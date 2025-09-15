import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys
import time

def print_step(step_number, message):
    """打印步骤提示"""
    time.sleep(0.5)
    print(f"\n=== {message} ===\n")

def wait_for_enter():
    """等待用户按回车继续"""
    input("按回车键继续...")


def find_header_row(sheet, header_keyword="种类"):
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        for cell_value in row:
            if cell_value and header_keyword in str(cell_value):
                print(f"检测到表头在第 {row_idx} 行")
                return row_idx
    print("未找到表头，默认返回第 1 行")
    return 1

def load_data(file_path):
    """加载Excel数据，只读取第1、3、6、8列"""
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # 查找表头行
    header_row = find_header_row(sheet)

    data = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        # 跳过空行
        if row[0] is None:
            continue

        row_data = {
            "种类": row[0],  # 第1列
            "工序1品质": row[2],  # 第3列
            "工序2品质": row[5],  # 第6列
            "最终品质": row[7]  # 第8列
        }
        data.append(row_data)

    return data


def preprocess_data(data):
    """预处理数据，填充最终品质的空值"""
    for item in data:
        # 如果最终品质为空，则根据工序2品质填充
        final_quality = item.get("最终品质")
        if final_quality is None or final_quality == "":
            process2_quality = item.get("工序2品质")
            if process2_quality == "好":
                item["最终品质"] = "好"
            else:
                item["最终品质"] = "坏"
    return data

## this is just an example

def quality_ratio(data, category, process_column, values_to_count, consider_empty=False):
    """
    计算指定种类、指定工序中特定值的占比
    参数:
    data: 数据集
    category: 产品种类
    process_column: 工序列名
    values_to_count: 要计算的值列表(如["好", "一般"])
    consider_empty: 是否考虑空值
    """
    total = 0
    count = 0

    for item in data:
        if item.get("种类") == category:
            value = item.get(process_column)

            if value is None or value == "":
                continue
            total += 1

            if value in values_to_count:
                count += 1
            # 如果考虑空值且值为空，则根据情况处理
            elif consider_empty and (value is None or value == ""):
                # 这里可以根据需要决定是否将空值计入特定类别
                pass

    return count / total if total > 0 else 0


def analyze_quality_data(quality_file_path, process_configs):
    """
    分析质量数据
    参数:
    file_path: Excel文件路径
    process_configs: 工序配置列表，每个配置是一个字典，包含:
        - name: 工序名称
        - column: 数据列名
        - values: 要计算的值列表
    """
    # 加载和预处理数据
    data = load_data(quality_file_path)
    if not data:
        print("未找到有效数据")
        return {}

    data = preprocess_data(data)

    # 获取所有种类
    categories = set()
    for item in data:
        category = item.get("种类")
        if category is not None:
            categories.add(category)

    results = {}

    for category in categories:
        category_results = {}

        # 为每个工序配置计算占比
        for config in process_configs:
            ratio = quality_ratio(data, category, config["column"], config["values"])
            category_results[config["name"]] = ratio

        results[category] = category_results

    return results

def save_results_to_excel(results, process_configs, output_file="质量分析结果.xlsx"):
    """将结果保存到Excel文件"""
    # 创建新的工作簿
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "质量分析结果"

    # 设置标题行
    headers = ["种类"]
    for config in process_configs:
        headers.append(config["name"])

    for col_idx, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    row_idx = 2
    for category, ratios in results.items():
        sheet.cell(row=row_idx, column=1, value=category)

        for col_idx, config in enumerate(process_configs, 2):
            sheet.cell(row=row_idx, column=col_idx, value=ratios[config["name"]]).number_format = '0.00%'

        row_idx += 1

    # 调整列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # 保存文件
    workbook.save(output_file)
    print(f"结果已保存到 {output_file}")


def print_results(results, process_configs):
    """打印结果"""
    for category, ratios in results.items():
        print(f"种类 {category}:")
        for config in process_configs:
            print(f"  {config['name']}: {ratios[config['name']]:.2%}")
        print()


# 主程序
if __name__ == "__main__":
    print_step(1, "长晶工艺质量数据分析")
    print("请确保Excel文件与此程序在同一文件夹中")
    wait_for_enter()

    # 查找Excel文件
    print_step(2, "查找Excel文件")
    excel_files = []
    for file in os.listdir('.'):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            excel_files.append(file)

    if not excel_files:
        print("未找到Excel文件(.xlsx或.xls)")
        print("请将Excel文件放入与此程序相同的文件夹中")
        input("按回车键退出...")
        sys.exit(1)

    # 如果找到多个Excel文件，让用户选择
    if len(excel_files) > 1:
        print_step(3, "发现多个Excel文件，请选择要分析的文件:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")

        while True:
            try:
                choice = int(input("请输入文件编号: "))
                if 1 <= choice <= len(excel_files):
                    file_path = excel_files[choice - 1]
                    break
                else:
                    print("编号无效，请重新输入")
            except ValueError:
                print("请输入有效数字")
    else:
        file_path = excel_files[0]
        print(f"找到文件: {file_path}")
        wait_for_enter()

    # 配置每个工序的参数
    process_configs = [
        {
            "name": "工序1_非坏占比",
            "column": "工序1品质",
            "values": ["好", "还行"]
        },
        {
            "name": "工序2_非坏占比",
            "column": "工序2品质",
            "values": ["好", "一般"]
        },
        {
            "name": "最终品质_好占比",
            "column": "最终品质",
            "values": ["好", "一般"]
        }
    ]

    print_step(4, "开始分析数据")
    print("正在读取和分析Excel文件...")

    try:
        # 分析数据
        results = analyze_quality_data(file_path, process_configs)

        print_step(5, "分析结果")
        # 打印结果
        print_results(results, process_configs)
        wait_for_enter()

        print_step(6, "保存结果")
        # 保存到Excel
        output_file = "质量分析结果.xlsx"
        save_results_to_excel(results, process_configs, output_file)

        print_step(7, "完成")
        print("所有操作已完成!")
        print("您可以在同一文件夹中找到 '质量分析结果.xlsx' 文件")

    except Exception as e:
        print(f"处理过程中发生错误: {str(e)}")
        import traceback

        traceback.print_exc()

    input("按回车键退出程序...")
